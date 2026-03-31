from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook


MAIN_WORKBOOK = 'Fluid table C10 Bacalhau_2026_03_26 (2).xlsx'
PVT_REFERENCE_WORKBOOK = 'pvt result.xlsx'


def _to_kgph(value_tph: float | None) -> float | None:
    if value_tph is None:
        return None
    return float(value_tph) * 1000.0


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    normalized = []
    for column in df.columns:
        text = unicodedata.normalize('NFKD', str(column)).encode('ascii', 'ignore').decode('ascii')
        text = text.strip().lower()
        text = re.sub(r'[^a-z0-9]+', '_', text).strip('_')
        normalized.append(text or 'column')
    result = df.copy()
    result.columns = normalized
    return result


def _match_column(columns: list[str], *patterns: str) -> str | None:
    for pattern in patterns:
        regex = re.compile(pattern)
        for column in columns:
            if regex.search(column):
                return column
    return None


def _to_numeric_column(table: pd.DataFrame, column: str | None) -> pd.Series:
    if column is None:
        return pd.Series([float('nan')] * len(table), index=table.index, dtype='float64')
    return pd.to_numeric(table[column], errors='coerce')


def _first_valid_numeric(table: pd.DataFrame, specs: list[tuple[str | None, float]]) -> pd.Series:
    result = pd.Series([float('nan')] * len(table), index=table.index, dtype='float64')
    for column, factor in specs:
        if column is None:
            continue
        candidate = _to_numeric_column(table, column) * factor
        result = result.where(result.notna(), candidate)
    return result


def _parse_generic_separator_table(df: pd.DataFrame) -> pd.DataFrame:
    table = _normalize_columns(df).dropna(how='all')
    columns = list(table.columns)

    case_id_col = _match_column(columns, r'^case_id$')
    case_label_col = _match_column(columns, r'^case_label$')
    separator_pressure_bara_col = _match_column(columns, r'^separator_pressure_bara$', r'^psep_bara$', r'separador.*pressao.*bara')
    separator_pressure_barg_col = _match_column(columns, r'^separator_pressure_barg$', r'^psep_barg$', r'separador.*pressao.*barg', r'referencia.*separador.*pressao.*barg')
    temperature_col = _match_column(columns, r'^separator_temperature_c$', r'^tsep_c$', r'(^|_)t(sep|emperature)(_c)?$', r'separador.*temperatura', r'referencia.*separador.*temperatura')
    oil_col = _match_column(
        columns,
        r'^separator_oil_kgph$',
        r'^separator_oil_kg$',
        r'^separator_oil_mass_kg$',
        r'separator.*oil.*kg',
        r'referencia.*separador.*oleo.*kg',
    )
    oil_t_col = _match_column(
        columns,
        r'^separator_oil_t$',
        r'^separator_oil_tonnes$',
        r'separador.*oleo.*_t($|_)',
        r'referencia.*separador.*oleo.*_t($|_)',
    )
    gas_col = _match_column(
        columns,
        r'^separator_gas_kgph$',
        r'^separator_gas_kg$',
        r'^separator_gas_mass_kg$',
        r'separator.*gas.*kg',
        r'referencia.*separador.*gas.*kg',
    )
    gas_t_col = _match_column(
        columns,
        r'^separator_gas_t$',
        r'^separator_gas_tonnes$',
        r'separador.*gas.*_t($|_)',
        r'referencia.*separador.*gas.*_t($|_)',
    )
    total_col = _match_column(columns, r'^separator_total_kgph$', r'^separator_total$', r'total')
    gor_col = _match_column(columns, r'^separator_gor_sm3_sm3$', r'^separator_gor$', r'gor', r'rgo')
    time_col = _match_column(columns, r'time', r'date', r'timestamp')
    separator_oil_volume_col = _match_column(
        columns,
        r'^separator_oil_volume_m3ph$',
        r'^separator_oil_volume_m3$',
        r'^separador_oleo_m3$',
        r'separador.*oleo.*m3',
        r'referencia.*separador.*oleo.*m3',
        r'oleo.*fonte_cv',
        r'volume.*oleo.*cv',
    )
    separator_oil_density_col = _match_column(
        columns,
        r'^separator_oil_density_kgm3$',
        r'^separator_density_kgm3$',
        r'^separador_densidade_oleo_kgm3$',
        r'density.*coriolis',
        r'densidade.*oleo',
        r'separador.*density',
    )
    mpfm_pressure_bara_col = _match_column(columns, r'^mpfm_pressure_bara$', r'^subsea_mpfm_pressure_bara$', r'^mpfm_pressao_bara$')
    mpfm_pressure_barg_col = _match_column(columns, r'^mpfm_pressure_barg$', r'^subsea_mpfm_pressure_barg$', r'^mpfm_pressure_barg$', r'^mpfm_pressao_barg$')
    mpfm_temperature_col = _match_column(columns, r'^mpfm_temperature_c$', r'^subsea_mpfm_temperature_c$', r'^mpfm_temperatura_c$')
    mpfm_oil_density_col = _match_column(columns, r'^mpfm_oil_density_kgm3$', r'^mpfm_densidade_oleo_kgm3$', r'^mpfm_densidade_oleo_kg_m3$')
    mpfm_gas_density_col = _match_column(columns, r'^mpfm_gas_density_kgm3$', r'^mpfm_densidade_gas_kgm3$', r'^mpfm_densidade_gas_kg_m3$')
    mpfm_oil_mass_col = _match_column(columns, r'^mpfm_oil_kgph$', r'^mpfm_subsea_oil_kgph$', r'mpfm.*subsea.*oleo.*kg')
    mpfm_oil_t_col = _match_column(columns, r'^mpfm_oil_t$', r'^mpfm_subsea_oil_t$', r'mpfm.*oleo.*_t$')
    mpfm_gas_mass_col = _match_column(columns, r'^mpfm_gas_kgph$', r'^mpfm_subsea_gas_kgph$', r'mpfm.*subsea.*gas.*kg')
    mpfm_gas_t_col = _match_column(columns, r'^mpfm_gas_t$', r'^mpfm_subsea_gas_t$', r'mpfm.*gas.*_t$')
    mpfm_water_mass_col = _match_column(columns, r'^mpfm_water_kgph$', r'^mpfm_subsea_water_kgph$', r'mpfm.*subsea.*agua.*kg', r'mpfm.*subsea.*water.*kg')
    mpfm_water_t_col = _match_column(columns, r'^mpfm_water_t$', r'^mpfm_subsea_water_t$', r'mpfm.*agua.*_t$', r'mpfm.*water.*_t$')
    fcs_oil_col = _match_column(columns, r'fcs.*oil', r'flow.*computer.*oil', r'model.*oil')
    fcs_gas_col = _match_column(columns, r'fcs.*gas', r'flow.*computer.*gas', r'model.*gas')
    fcs_gor_col = _match_column(columns, r'fcs.*gor', r'flow.*computer.*gor', r'model.*gor')

    result = pd.DataFrame(index=table.index)
    if case_id_col is not None:
        result['case_id'] = table[case_id_col].astype(str)
    elif time_col is not None:
        result['case_id'] = pd.to_datetime(table[time_col], errors='coerce').astype(str)
    else:
        result['case_id'] = [f'case_{index + 1:03d}' for index in range(len(table))]

    if case_label_col is not None:
        result['case_label'] = table[case_label_col].astype(str)
    else:
        result['case_label'] = result['case_id']
    separator_bara = _to_numeric_column(table, separator_pressure_bara_col)
    if separator_bara.isna().all() and separator_pressure_barg_col is not None:
        separator_bara = _to_numeric_column(table, separator_pressure_barg_col) + 1.01325
    result['separator_pressure_bara'] = separator_bara
    result['separator_temperature_c'] = _to_numeric_column(table, temperature_col)
    result['separator_oil_kgph'] = _first_valid_numeric(table, [(oil_col, 1.0), (oil_t_col, 1000.0)])
    result['separator_gas_kgph'] = _first_valid_numeric(table, [(gas_col, 1.0), (gas_t_col, 1000.0)])
    result['separator_total_kgph'] = _to_numeric_column(table, total_col)
    result['separator_gor_sm3_sm3'] = _to_numeric_column(table, gor_col)
    result['separator_oil_volume_m3ph'] = _to_numeric_column(table, separator_oil_volume_col)
    result['separator_oil_density_kgm3'] = _to_numeric_column(table, separator_oil_density_col)

    mpfm_bara = _to_numeric_column(table, mpfm_pressure_bara_col)
    if mpfm_bara.isna().all() and mpfm_pressure_barg_col is not None:
        mpfm_bara = _to_numeric_column(table, mpfm_pressure_barg_col) + 1.01325
    result['mpfm_pressure_bara'] = mpfm_bara
    result['mpfm_temperature_c'] = _to_numeric_column(table, mpfm_temperature_col)
    result['mpfm_oil_density_kgm3'] = _to_numeric_column(table, mpfm_oil_density_col)
    result['mpfm_gas_density_kgm3'] = _to_numeric_column(table, mpfm_gas_density_col)
    result['mpfm_oil_kgph'] = _first_valid_numeric(table, [(mpfm_oil_mass_col, 1.0), (mpfm_oil_t_col, 1000.0)])
    result['mpfm_gas_kgph'] = _first_valid_numeric(table, [(mpfm_gas_mass_col, 1.0), (mpfm_gas_t_col, 1000.0)])
    result['mpfm_water_kgph'] = _first_valid_numeric(table, [(mpfm_water_mass_col, 1.0), (mpfm_water_t_col, 1000.0)])

    result['fcs320_oil_kgph'] = _to_numeric_column(table, fcs_oil_col)
    result['fcs320_gas_kgph'] = _to_numeric_column(table, fcs_gas_col)
    result['fcs320_gor_sm3_sm3'] = _to_numeric_column(table, fcs_gor_col)
    result['source_name'] = 'uploaded_table'

    if result['separator_total_kgph'].isna().all():
        result['separator_total_kgph'] = result['separator_oil_kgph'].fillna(0.0) + result['separator_gas_kgph'].fillna(0.0)
    return result.dropna(subset=['separator_oil_kgph', 'separator_gas_kgph'], how='all')


def load_built_in_separator_cases(base_dir: str | Path) -> pd.DataFrame:
    workbook_path = Path(base_dir) / MAIN_WORKBOOK
    reference_path = Path(base_dir) / PVT_REFERENCE_WORKBOOK

    rows: list[dict[str, object]] = []

    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    ws_oct = wb['2025.10 Sep Test Sample']
    row_p = next(ws_oct.iter_rows(min_row=3, max_row=3, values_only=True))
    row_t = next(ws_oct.iter_rows(min_row=4, max_row=4, values_only=True))
    row_oil = next(ws_oct.iter_rows(min_row=5, max_row=5, values_only=True))
    row_gas = next(ws_oct.iter_rows(min_row=6, max_row=6, values_only=True))
    row_gor = next(ws_oct.iter_rows(min_row=8, max_row=8, values_only=True))

    rows.extend(
        [
            {
                'case_id': 'oct2025_hot_sample018',
                'case_label': 'Oct 2025 | 64.45 C | sample 018+017',
                'separator_pressure_bara': float(row_p[2]),
                'separator_temperature_c': float(row_t[2]),
                'separator_oil_kgph': _to_kgph(float(row_oil[2])),
                'separator_gas_kgph': _to_kgph(float(row_gas[2])),
                'separator_total_kgph': _to_kgph(float(row_oil[2]) + float(row_gas[2])),
                'separator_gor_sm3_sm3': float(row_gor[3]),
                'fcs320_oil_kgph': None,
                'fcs320_gas_kgph': None,
                'fcs320_gor_sm3_sm3': float(row_gor[4]),
                'source_name': 'built_in',
            },
            {
                'case_id': 'oct2025_hot_sample015',
                'case_label': 'Oct 2025 | 64.45 C | sample 015+017',
                'separator_pressure_bara': float(row_p[2]),
                'separator_temperature_c': float(row_t[2]),
                'separator_oil_kgph': _to_kgph(float(row_oil[2])),
                'separator_gas_kgph': _to_kgph(float(row_gas[2])),
                'separator_total_kgph': _to_kgph(float(row_oil[2]) + float(row_gas[2])),
                'separator_gor_sm3_sm3': float(row_gor[6]),
                'fcs320_oil_kgph': None,
                'fcs320_gas_kgph': None,
                'fcs320_gor_sm3_sm3': float(row_gor[7]),
                'source_name': 'built_in',
            },
            {
                'case_id': 'oct2025_cold_sample018',
                'case_label': 'Oct 2025 | 29.90 C | sample 018+017',
                'separator_pressure_bara': float(row_p[11]),
                'separator_temperature_c': float(row_t[11]),
                'separator_oil_kgph': _to_kgph(float(row_oil[11])),
                'separator_gas_kgph': _to_kgph(float(row_gas[11])),
                'separator_total_kgph': _to_kgph(float(row_oil[11]) + float(row_gas[11])),
                'separator_gor_sm3_sm3': float(row_gor[12]),
                'fcs320_oil_kgph': None,
                'fcs320_gas_kgph': None,
                'fcs320_gor_sm3_sm3': float(row_gor[13]),
                'source_name': 'built_in',
            },
            {
                'case_id': 'oct2025_cold_sample015',
                'case_label': 'Oct 2025 | 29.90 C | sample 015+017',
                'separator_pressure_bara': float(row_p[11]),
                'separator_temperature_c': float(row_t[11]),
                'separator_oil_kgph': _to_kgph(float(row_oil[11])),
                'separator_gas_kgph': _to_kgph(float(row_gas[11])),
                'separator_total_kgph': _to_kgph(float(row_oil[11]) + float(row_gas[11])),
                'separator_gor_sm3_sm3': float(row_gor[15]),
                'fcs320_oil_kgph': None,
                'fcs320_gas_kgph': None,
                'fcs320_gor_sm3_sm3': float(row_gor[16]),
                'source_name': 'built_in',
            },
        ]
    )

    ws_feb = wb['2026.02 Sep Test Sample PE_2']
    p_row = next(ws_feb.iter_rows(min_row=3, max_row=3, values_only=True))
    t_row = next(ws_feb.iter_rows(min_row=4, max_row=4, values_only=True))
    oil_row = next(ws_feb.iter_rows(min_row=5, max_row=5, values_only=True))
    gas_row = next(ws_feb.iter_rows(min_row=6, max_row=6, values_only=True))
    gor_row = next(ws_feb.iter_rows(min_row=10, max_row=10, values_only=True))
    rows.append(
        {
            'case_id': 'feb2026_pe2',
            'case_label': 'Feb 2026 | PE-2',
            'separator_pressure_bara': float(p_row[2]),
            'separator_temperature_c': float(t_row[2]),
            'separator_oil_kgph': _to_kgph(float(oil_row[2])),
            'separator_gas_kgph': _to_kgph(float(gas_row[2])),
            'separator_total_kgph': _to_kgph(float(oil_row[2]) + float(gas_row[2])),
            'separator_gor_sm3_sm3': float(gor_row[2]),
            'fcs320_oil_kgph': None,
            'fcs320_gas_kgph': None,
            'fcs320_gor_sm3_sm3': float(gor_row[3]),
            'source_name': 'built_in',
        }
    )

    ref_wb = load_workbook(reference_path, read_only=True, data_only=True)
    ref_ws = ref_wb['2-Valores Calculados']
    rows.append(
        {
            'case_id': 'calibration_reference_oct2025',
            'case_label': 'Calibration baseline | Oct 2025',
            'separator_pressure_bara': None,
            'separator_temperature_c': None,
            'separator_oil_kgph': float(next(ref_ws.iter_rows(min_row=25, max_row=25, values_only=True))[1]),
            'separator_gas_kgph': float(next(ref_ws.iter_rows(min_row=26, max_row=26, values_only=True))[1]),
            'separator_total_kgph': float(next(ref_ws.iter_rows(min_row=24, max_row=24, values_only=True))[1]),
            'separator_gor_sm3_sm3': None,
            'fcs320_oil_kgph': float(next(ref_ws.iter_rows(min_row=25, max_row=25, values_only=True))[2]),
            'fcs320_gas_kgph': float(next(ref_ws.iter_rows(min_row=26, max_row=26, values_only=True))[2]),
            'fcs320_gor_sm3_sm3': None,
            'source_name': 'pvt_result',
        }
    )

    result = pd.DataFrame(rows)
    for column in [
        'separator_oil_volume_m3ph',
        'separator_oil_density_kgm3',
        'mpfm_pressure_bara',
        'mpfm_temperature_c',
        'mpfm_oil_density_kgm3',
        'mpfm_gas_density_kgm3',
        'mpfm_oil_kgph',
        'mpfm_gas_kgph',
        'mpfm_water_kgph',
    ]:
        if column not in result.columns:
            result[column] = float('nan')
    return result


def load_separator_upload(uploaded_file) -> pd.DataFrame:
    suffix = Path(uploaded_file.name).suffix.lower()
    raw = uploaded_file.getvalue()

    if suffix == '.csv':
        return _parse_generic_separator_table(pd.read_csv(BytesIO(raw)))

    if suffix in {'.xlsx', '.xlsm', '.xls'}:
        excel = pd.ExcelFile(BytesIO(raw))
        candidate_sheet = excel.sheet_names[0]
        for sheet_name in excel.sheet_names:
            parsed = _parse_generic_separator_table(excel.parse(sheet_name))
            if not parsed.empty:
                return parsed
            normalized_columns = set(_normalize_columns(excel.parse(sheet_name).head(0)).columns)
            if {'separator_pressure_bara', 'separator_pressure_barg', 'separator_oil_kgph', 'separator_oil_t'} & normalized_columns:
                candidate_sheet = sheet_name
        return _parse_generic_separator_table(excel.parse(candidate_sheet))

    raise ValueError(f'Unsupported file format: {suffix}')
