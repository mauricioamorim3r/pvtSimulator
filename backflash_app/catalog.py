from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .models import CompositionScenario


MAIN_WORKBOOK = 'Fluid table C10 Bacalhau_2026_03_26 (2).xlsx'


def _clean_text(value: object) -> str:
    return str(value).strip() if value is not None else ''


def _build_scenario(
    key: str,
    label: str,
    source: str,
    estimated_gor: float | None,
    components: dict[str, float],
    plus_mw: float | None,
    plus_density: float | None,
    **metadata: object,
) -> CompositionScenario:
    return CompositionScenario(
        key=key,
        label=label,
        source=source,
        estimated_gor_sm3_sm3=estimated_gor,
        components_mol_pct=components,
        plus_mw_g_mol=plus_mw,
        plus_density_g_cc=plus_density,
        metadata={k: v for k, v in metadata.items() if v is not None},
    )


def _parse_main_composition_sheet(workbook_path: Path) -> dict[str, CompositionScenario]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb['Composition_table']

    gor_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    name_row = next(ws.iter_rows(min_row=5, max_row=5, values_only=True))

    scenario_columns: list[tuple[int, str, float | None]] = []
    for idx, value in enumerate(name_row):
        text = _clean_text(value)
        if text.startswith('GOR_'):
            estimated_gor = gor_row[idx] if idx < len(gor_row) else None
            scenario_columns.append((idx, text, float(estimated_gor) if estimated_gor is not None else None))

    component_rows: dict[str, tuple] = {}
    for row in ws.iter_rows(min_row=8, max_row=24, values_only=True):
        name = _clean_text(row[1] if len(row) > 1 else None)
        if name:
            component_rows[name] = row

    scenarios: dict[str, CompositionScenario] = {}
    for column_index, scenario_name, estimated_gor in scenario_columns:
        components: dict[str, float] = {}
        for component in [
            'N2',
            'CO2',
            'C1',
            'C2',
            'C3',
            'I-C4',
            'N-C4',
            'I-C5',
            'N-C5',
            'C6',
            'C7',
            'C8',
            'C9',
            'C10+',
        ]:
            row = component_rows.get(component)
            if row is not None and column_index < len(row) and row[column_index] is not None:
                components[component] = float(row[column_index])

        plus_mw_row = component_rows.get('C10+ mol wgt')
        plus_density_row = component_rows.get('C10+ Density (gr/cc)')
        plus_mw = float(plus_mw_row[column_index]) if plus_mw_row and plus_mw_row[column_index] is not None else None
        plus_density = (
            float(plus_density_row[column_index])
            if plus_density_row and plus_density_row[column_index] is not None
            else None
        )

        scenarios[scenario_name] = _build_scenario(
            key=scenario_name,
            label=f'{scenario_name} | main catalog',
            source='composition_table',
            estimated_gor=estimated_gor,
            components=components,
            plus_mw=plus_mw,
            plus_density=plus_density,
        )
    return scenarios


def _parse_october_2025_sheet(workbook_path: Path) -> dict[str, CompositionScenario]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb['2025.10 Sep Test Sample']

    component_names = ['N2', 'CO2', 'C1', 'C2', 'C3', 'I-C4', 'N-C4', 'I-C5', 'N-C5', 'C6', 'C7', 'C8', 'C9', 'C10+']
    components_a: dict[str, float] = {}
    components_b: dict[str, float] = {}

    for row_index, component in enumerate(component_names, start=11):
        row = next(ws.iter_rows(min_row=row_index, max_row=row_index, values_only=True))
        components_a[component] = float(row[4])
        components_b[component] = float(row[7])

    plus_mw_row = next(ws.iter_rows(min_row=26, max_row=26, values_only=True))
    plus_density_row = next(ws.iter_rows(min_row=27, max_row=27, values_only=True))
    gor_row_hot = next(ws.iter_rows(min_row=8, max_row=8, values_only=True))

    return {
        'oct2025_sample_018_017': _build_scenario(
            key='oct2025_sample_018_017',
            label='oct2025 | oil018 + gas017',
            source='2025.10 Sep Test Sample',
            estimated_gor=float(gor_row_hot[4]),
            components=components_a,
            plus_mw=float(plus_mw_row[4]),
            plus_density=float(plus_density_row[4]),
            sample='018+017',
        ),
        'oct2025_sample_015_017': _build_scenario(
            key='oct2025_sample_015_017',
            label='oct2025 | oil015 + gas017',
            source='2025.10 Sep Test Sample',
            estimated_gor=float(gor_row_hot[7]),
            components=components_b,
            plus_mw=float(plus_mw_row[7]),
            plus_density=float(plus_density_row[7]),
            sample='015+017',
        ),
    }


def _parse_february_2026_sheet(workbook_path: Path) -> dict[str, CompositionScenario]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb['2026.02 Sep Test Sample PE_2']

    components: dict[str, float] = {}
    for row_index, component in enumerate(
        ['N2', 'CO2', 'C1', 'C2', 'C3', 'I-C4', 'N-C4', 'I-C5', 'N-C5', 'C6', 'C7', 'C8', 'C9', 'C10+'],
        start=14,
    ):
        row = next(ws.iter_rows(min_row=row_index, max_row=row_index, values_only=True))
        components[component] = float(row[2])

    gor_row = next(ws.iter_rows(min_row=10, max_row=10, values_only=True))
    plus_mw_row = next(ws.iter_rows(min_row=29, max_row=29, values_only=True))
    plus_density_row = next(ws.iter_rows(min_row=30, max_row=30, values_only=True))
    return {
        'feb2026_pe2': _build_scenario(
            key='feb2026_pe2',
            label='feb2026 | PE-2 recombined',
            source='2026.02 Sep Test Sample PE_2',
            estimated_gor=float(gor_row[2]),
            components=components,
            plus_mw=float(plus_mw_row[2]),
            plus_density=float(plus_density_row[2]),
        )
    }


def load_composition_catalog(base_dir: str | Path) -> dict[str, CompositionScenario]:
    workbook_path = Path(base_dir) / MAIN_WORKBOOK
    scenarios: dict[str, CompositionScenario] = {}
    scenarios.update(_parse_main_composition_sheet(workbook_path))
    scenarios.update(_parse_october_2025_sheet(workbook_path))
    scenarios.update(_parse_february_2026_sheet(workbook_path))
    return scenarios
